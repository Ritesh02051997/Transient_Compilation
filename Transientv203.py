import mdfreader
import openpyxl
from os import listdir,path,makedirs
from tkinter import *
from openpyxl.chart import Reference, Series,ScatterChart
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.formatting.rule import CellIsRule
from shutil import move,copy

def transfer_files(current_file,dest_path, folder_action) :
    source_dat = current_file
    if folder_action == 'Copy':
        copy(source_dat, dest_path)
    else:
        move(source_dat, dest_path)

def compile(data,ini_path,folder_action,template_file,alias_manager) :
    for count in range(len(data['file'])) :
        raster = 0.01
        mdfData=mdfreader.Mdf(data['file'][count])
        mdfData.resample(raster)
        #For fuel_check
        if mdfData.get_channel_data('B_cng') is not None :
            B_cng_array = mdfData.get_channel_data('B_cng')
            B_cng = sum(B_cng_array)/len(B_cng_array)
            if B_cng > 0 :
                Fuel = 'CNG'
            else :
                Fuel = 'Gasoline'
        else:
            Fuel = 'Gasoline'
        #For all Alias_names
        alias_excel = openpyxl.load_workbook(alias_manager)
        alias_sheet = alias_excel["Alias_Names"]
        len_a_rpm = [i for i in range(1,100) if alias_sheet['B'+str(i)].value is not None]
        len_a_rl = [i for i in range(1,100) if alias_sheet['C'+str(i)].value is not None]
        len_a_lambda = [i for i in range(1,100) if alias_sheet['D'+str(i)].value is not None]
        len_a_tmot = [i for i in range(1,100) if alias_sheet['E'+str(i)].value is not None]
        len_a_app = [i for i in range(1,100) if alias_sheet['F'+str(i)].value is not None]
        len_a_load = [i for i in range(1,100) if alias_sheet['G'+str(i)].value is not None]
        rpm = 0
        for i in range(2,len_a_rpm[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['B'+str(i)].value)) is not None :
                rpm_array = mdfData.get_channel_data(str(alias_sheet['B'+str(i)].value))
                rpm = int(round((sum(rpm_array)/len(rpm_array)),-2))
                break
        for i in range(2,len_a_rl[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['C'+str(i)].value)) is not None :
                rlsol = alias_sheet['C'+str(i)].value
                break
        for i in range(2,len_a_lambda[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['D'+str(i)].value)) is not None :
                lambda_name = alias_sheet['D'+str(i)].value
                break
        for i in range(2,len_a_tmot[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['E'+str(i)].value)) is not None :
                tmot = alias_sheet['E'+str(i)].value
                break
        for i in range(2,len_a_app[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['F'+str(i)].value)) is not None :
                app = alias_sheet['F'+str(i)].value
                break
        for i in range(2,len_a_app[-1]+1) :
            if mdfData.get_channel_data(str(alias_sheet['G'+str(i)].value)) is not None :
                load = alias_sheet['G'+str(i)].value
                break
        #Main Calculations
        mdfData.cut(master_channel = 'time', begin = float(data['start_time'][count]), end = float(data['end_time'][count]))
        if data['method'][count] == 'APP_r' :
            splitting_var_data = mdfData.get_channel_data(app)
        elif data['method'][count] == 'RLSOLAP' :
            splitting_var_data = mdfData.get_channel_data(rlsol)
        time = float(data['end_time'][count])
        time_splits = [ 0 ]
        rl_array = []
        temp_split_var = splitting_var_data[0]
        counter = 0
        if data['method'][count] == 'APP_r' :
            for index in range(1,len(splitting_var_data)) :
                if counter > 300 :
                    if not (splitting_var_data[index]-1.5)<temp_split_var<(splitting_var_data[index]+1.5) :
                        time_splits.append(index*raster)
                        temp_split_var = splitting_var_data[index+300]
                        counter = 0
                else :
                    counter = counter + 1
        elif data['method'][count] == 'RLSOLAP' :
            for index in range(1,len(splitting_var_data)) :
                if counter > 300 :
                    if not (splitting_var_data[index]-5)<temp_split_var<(splitting_var_data[index]+5) :
                        time_splits.append(index*raster)
                        temp_split_var = splitting_var_data[index+300]
                        counter = 0
                else :
                    counter = counter + 1
        time_splits.append(float(data['end_time'][count]))
        load_data = mdfData.get_channel_data(load)
        lambda_array = []
        splitting_var_array = []
        splitting_var_array_mean = []
        rl_array_time = []

        rl_array_mean = []
        min_lambda_a = []
        max_lambda_a = []
        min_lambda_d = []
        max_lambda_d = []
        temperature = []
        mean_lambda_array = []
        exact_time=[]
        for i in range(len(time_splits)):
            exact_time.append(time_splits[i]+float(data['start_time'][count]))
        lambda_data = mdfData.get_channel_data(lambda_name)
        temperature_data = mdfData.get_channel_data(tmot)
        deg = int(round(temperature_data[1],-1))
        if temperature_data[1] > deg :
            if (temperature_data[1] - deg)>3:
                deg = deg + 5
        else :
            if (deg - temperature_data[1])>3:
                deg = deg - 5

        dest_path = path.join(ini_path,str(data['project'][count]), Fuel, str(deg)+' Deg')
        if not path.exists(dest_path):
            makedirs(dest_path)
        output_file_name = str(str(dest_path)+'/'+str(deg)+'DegC Transient.xlsx')
        if path.exists(output_file_name) :
            output_workbook = openpyxl.load_workbook(output_file_name)
        else :
            output_workbook = openpyxl.load_workbook(template_file)
        tab_name = str(str(deg)+'DegC_'+str(rpm)+'RPM')
        #Splitting
        for i in range(len(time_splits)-1) :
            lambda_array.append( lambda_data[int(time_splits[i]/raster):int(time_splits[i+1]/raster)] )
            splitting_var_array.append( splitting_var_data[int(time_splits[i]/raster):int(time_splits[i+1]/raster)] )
            splitting_var_array_mean.append(sum(splitting_var_array[-1])/len(splitting_var_array[-1]))
            rl_array_time.append(load_data[int(time_splits[i]/raster):int(time_splits[i+1]/raster)] )
            rl_array_mean.append(sum(rl_array_time[-1])/len(rl_array_time[-1]))
            temperature.append(temperature_data[int(time_splits[i]/raster)])
            len_lambda_array = len(lambda_array[-1])
            min_lambda_a.append(min(lambda_array[-1][int(0/raster):int(3/raster)]))
            max_lambda_a.append(max(lambda_array[-1][int(0/raster):int(3/raster)]))

            min_lambda_d.append(min(lambda_array[-1][int(0.2/raster):int(3/raster)]))
            max_lambda_d.append(max(lambda_array[-1][int(0.2/raster):int(3/raster)]))

            mean_lambda_array.append(sum([lambda_array[-1][i] for i in range(round(0.8*len_lambda_array),len_lambda_array)])/(len_lambda_array-round(0.8*len_lambda_array)))


        #Writing
        worksheet = output_workbook.copy_worksheet(output_workbook['Template'])
        worksheet.title = path.basename(tab_name)
        worksheet.cell(row = 2, column = 2).value = path.basename(tab_name)
        index_a = 1
        index_d = 1
        #Acceleration
        for row in range(1,len(lambda_array)):
            if (round(rl_array_mean[row]) > round(rl_array_mean[row-1])) :
                worksheet.cell(row = index_a + 4, column = 2).value = index_a
                worksheet.cell(row = index_a + 4, column = 3).value = round(exact_time[row],2)
                worksheet.cell(row = index_a + 4, column = 4).value = round(temperature[row],2)
                worksheet.cell(row = index_a + 4, column = 5).value = round(rl_array_mean[row-1],0)
                worksheet.cell(row = index_a + 4, column = 6).value = round(rl_array_mean[row],0)

                worksheet.cell(row = index_a + 4, column = 7).value = round(mean_lambda_array[row-1],2)
                if (max_lambda_a[row]<=mean_lambda_array[row]) :
                    max_lambda_a[row]=mean_lambda_array[row]
                worksheet.cell(row = index_a + 4, column = 8).value = round(max_lambda_a[row],2)
                worksheet.cell(row = index_a + 4, column = 9).value = round((max_lambda_a[row] - mean_lambda_array[row]),2)
                if (min_lambda_a[row]>=mean_lambda_array[row-1]) :
                    min_lambda_a[row]=mean_lambda_array[row-1]
                worksheet.cell(row = index_a + 4, column = 10).value = round(min_lambda_a[row],2)
                worksheet.cell(row = index_a + 4, column = 11).value = round(mean_lambda_array[row],2)
                worksheet.cell(row = index_a + 4, column = 12).value = round((min_lambda_a[row] - mean_lambda_array[row-1]),2)
                index_a = index_a + 1

            #Deceleration
            if (round(rl_array_mean[row]) < round(rl_array_mean[row-1])) :
                worksheet.cell(row = index_d + 4, column = 14).value = index_d
                worksheet.cell(row = index_d + 4, column = 15).value = round(exact_time[row],2)
                worksheet.cell(row = index_d + 4, column = 16).value = round(temperature[row],2)
                worksheet.cell(row = index_d + 4, column = 17).value = round(rl_array_mean[row-1],0)
                worksheet.cell(row = index_d + 4, column = 18).value = round(rl_array_mean[row],0)

                worksheet.cell(row = index_d + 4, column = 19).value = round(mean_lambda_array[row-1],2)
                if (max_lambda_d[row]<=mean_lambda_array[row]) :
                    max_lambda_d[row]=mean_lambda_array[row]
                worksheet.cell(row = index_d + 4, column = 20).value = round(max_lambda_d[row],2)
                worksheet.cell(row = index_d + 4, column = 21).value = round((max_lambda_d[row] - mean_lambda_array[row]),2)
                if (min_lambda_d[row]>=mean_lambda_array[row-1]) :
                    min_lambda_d[row]=mean_lambda_array[row-1]
                worksheet.cell(row = index_d + 4, column = 22).value = round(min_lambda_d[row],2)
                worksheet.cell(row = index_d + 4, column = 23).value = round(mean_lambda_array[row],2)
                worksheet.cell(row = index_d + 4, column = 24).value = round((min_lambda_d[row] - mean_lambda_array[row-1]),2)
                index_d = index_d + 1

        #Conditional Formatting
        def cond(up_a,up_d,down_a,down_d) :
            redFill = PatternFill(start_color='F75D59', end_color='F75D59', fill_type='solid')
            worksheet.conditional_formatting.add('I5:I'+str(index_a+4),CellIsRule(operator='greaterThan', formula=[(up_a/100)], fill=redFill))
            worksheet.conditional_formatting.add('L5:L'+str(index_a+4),CellIsRule(operator='lessThan', formula=[-(down_a/100)], fill=redFill))
            worksheet.conditional_formatting.add('U5:U'+str(index_a+4),CellIsRule(operator='greaterThan', formula=[(up_d/100)], fill=redFill))
            worksheet.conditional_formatting.add('X5:X'+str(index_a+4),CellIsRule(operator='lessThan', formula=[-(down_d/100)], fill=redFill))

        up = str(data['up'][count]).split(',')
        down = str(data['down'][count]).split(',')
        if len(up) is 1 :
            up_d = up_a = int(up[0])
        elif len(up) is 2 :
            up_a = int(up[0])
            up_d = int(up[1])
        if len(down) is 1 :
            down_d = down_a = int(down[0])
        elif len(down) is 2:
            down_a = int(down[0])
            down_d = int(down[1])
        cond(up_a,up_d,down_a,down_d)


        index_f = (index_a if (index_a >= index_d) else index_d) + 3
        #For Graph of acceleration Portion
        A_axis = Reference(worksheet, min_col = 4, min_row = 5, max_col = 4, max_row = index_f)
        D_axis = Reference(worksheet, min_col = 16, min_row = 5, max_col = 16, max_row = index_f)
        A_Dev_Lean = Reference(worksheet, min_col = 9, min_row = 5, max_col = 9, max_row = index_f)
        A_Dev_Rich = Reference(worksheet, min_col = 12, min_row = 5, max_col = 12, max_row = index_f)
        D_Dev_Lean = Reference(worksheet, min_col = 21, min_row = 5, max_col = 21, max_row = index_f)
        D_Dev_Rich = Reference(worksheet, min_col = 24, min_row = 5, max_col = 24, max_row = index_f)
        A_chart = ScatterChart()
        A_chart.height = 10
        A_chart.width = 18
        D_chart = ScatterChart()
        D_chart.height = 10
        D_chart.width = 18
        Dev_Lean_a = Series(values = A_Dev_Lean, xvalues = A_axis, title ="Lean")
        Dev_Rich_a = Series(values = A_Dev_Rich, xvalues = A_axis, title ="Rich")
        A_chart.series.append(Dev_Lean_a)
        A_chart.series.append(Dev_Rich_a)
        A_chart.title = path.basename(tab_name)
        A_chart.x_axis.title = " Test No. "
        A_chart.y_axis.title = " λ "
        #For Grapgh at decelration portion
        Dev_Lean_a.marker=openpyxl.chart.marker.Marker('triangle', size = '7')
        Dev_Rich_a.marker=openpyxl.chart.marker.Marker('triangle', size = '7')
        Dev_Lean_a.graphicalProperties.line.noFill=True
        Dev_Rich_a.graphicalProperties.line.noFill=True
        worksheet.add_chart(A_chart, "B"+str(index_f+2))
        Dev_Lean_d = Series(values = D_Dev_Lean, xvalues = D_axis, title ="Lean")
        Dev_Rich_d = Series(values = D_Dev_Rich, xvalues = D_axis, title ="Rich")
        D_chart.series.append(Dev_Lean_d)
        D_chart.series.append(Dev_Rich_d)
        D_chart.title = path.basename(tab_name)
        D_chart.x_axis.title = " Test No. "
        D_chart.y_axis.title = " λ "
        Dev_Lean_d.marker=openpyxl.chart.marker.Marker('triangle', size = '7')
        Dev_Rich_d.marker=openpyxl.chart.marker.Marker('triangle', size = '7')
        Dev_Lean_d.graphicalProperties.line.noFill=True
        Dev_Rich_d.graphicalProperties.line.noFill=True
        worksheet.add_chart(D_chart, "N"+str(index_f+2))
        output_workbook.save(output_file_name)
        transfer_files(data['file'][count],dest_path,folder_action)
        print('Done Writing')
