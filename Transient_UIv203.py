# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Transient.ui'
#
# Created by: PyQt5 UI code generator 5.14.1
#
# WARNING! All changes made in this file will be lost!


from PyQt5 import QtCore, QtGui, QtWidgets
from pyqtgraph import PlotWidget, plot, mkPen
import pyqtgraph as pg
import os
import openpyxl
import mdfreader
import Transientv203
from random import choice
class Ui_Transient_Compilation(object):
    def setupUi(self, Transient_Compilation):
        Transient_Compilation.setObjectName("Transient_Compilation")
        Transient_Compilation.resize(1260, 850)
        self.centralwidget = QtWidgets.QWidget(Transient_Compilation)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setObjectName("verticalLayout")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setVerticalSpacing(15)
        self.gridLayout.setObjectName("gridLayout")
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_16 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_16.sizePolicy().hasHeightForWidth())
        self.label_16.setSizePolicy(sizePolicy)
        self.label_16.setObjectName("label_16")
        self.label_16.setFont(font)
        self.gridLayout.addWidget(self.label_16, 1, 0, 1, 1)
        self.button_link = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_link.setFont(font)
        self.button_link.setObjectName("button_link")
        self.gridLayout.addWidget(self.button_link, 3, 10, 1, 1)
        self.line_rlsol = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_rlsol.sizePolicy().hasHeightForWidth())
        self.line_rlsol.setSizePolicy(sizePolicy)
        self.line_rlsol.setObjectName("line_rlsol")
        self.gridLayout.addWidget(self.line_rlsol, 2, 10, 1, 1)
        self.rb_move = QtWidgets.QRadioButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.rb_move.sizePolicy().hasHeightForWidth())
        self.rb_move.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.rb_move.setFont(font)
        self.rb_move.setObjectName("rb_move")
        self.buttonGroup = QtWidgets.QButtonGroup(Transient_Compilation)
        self.buttonGroup.setObjectName("buttonGroup")
        self.buttonGroup.addButton(self.rb_move)
        self.gridLayout.addWidget(self.rb_move, 3, 8, 1, 1)
        self.rb_copy = QtWidgets.QRadioButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.rb_copy.setFont(font)
        self.rb_copy.setObjectName("rb_copy")
        self.buttonGroup.addButton(self.rb_copy)
        self.gridLayout.addWidget(self.rb_copy, 3, 9, 1, 1)
        self.button_output_folder = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_output_folder.setFont(font)
        self.button_output_folder.setObjectName("button_output_folder")
        self.gridLayout.addWidget(self.button_output_folder, 4, 10, 1, 1)
        self.label = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label.sizePolicy().hasHeightForWidth())
        self.label.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 1, 2, 1, 1)
        self.line_pn = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_pn.sizePolicy().hasHeightForWidth())
        self.line_pn.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.line_pn.setFont(font)
        self.line_pn.setObjectName("line_pn")
        self.gridLayout.addWidget(self.line_pn, 1, 3, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_2.sizePolicy().hasHeightForWidth())
        self.label_2.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 1, 4, 1, 1)
        self.line_start = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_start.sizePolicy().hasHeightForWidth())
        self.line_start.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.line_start.setFont(font)
        self.line_start.setObjectName("line_start")
        self.gridLayout.addWidget(self.line_start, 1, 5, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_3.sizePolicy().hasHeightForWidth())
        self.label_3.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.gridLayout.addWidget(self.label_3, 1, 6, 1, 1)
        self.drop_lambda = QtWidgets.QComboBox(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.drop_lambda.sizePolicy().hasHeightForWidth())
        self.drop_lambda.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.drop_lambda.setFont(font)
        self.drop_lambda.setObjectName("drop_lambda")
        self.gridLayout.addWidget(self.drop_lambda, 1, 9, 1, 1)
        self.button_lambda = QtWidgets.QPushButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.button_lambda.sizePolicy().hasHeightForWidth())
        self.button_lambda.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_lambda.setFont(font)
        self.button_lambda.setObjectName("button_lambda")
        self.gridLayout.addWidget(self.button_lambda, 1, 11, 1, 1)
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.gridLayout.addWidget(self.label_6, 2, 4, 1, 1)
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_4.sizePolicy().hasHeightForWidth())
        self.label_4.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_4.setObjectName("label_4")
        self.gridLayout.addWidget(self.label_4, 1, 8, 1, 1)
        self.drop_rlsol = QtWidgets.QComboBox(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.drop_rlsol.sizePolicy().hasHeightForWidth())
        self.drop_rlsol.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.drop_rlsol.setFont(font)
        self.drop_rlsol.setObjectName("drop_rlsol")
        self.gridLayout.addWidget(self.drop_rlsol, 2, 9, 1, 1)
        self.button_reset = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_reset.setFont(font)
        self.button_reset.setObjectName("button_reset")
        self.gridLayout.addWidget(self.button_reset, 4, 8, 1, 1)
        self.button_compile = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_compile.setFont(font)
        self.button_compile.setObjectName("button_compile")
        self.gridLayout.addWidget(self.button_compile, 4, 9, 1, 1)
        self.button_dat = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_dat.setFont(font)
        self.button_dat.setObjectName("button_dat")
        self.gridLayout.addWidget(self.button_dat, 3, 6, 1, 1)
        self.line_end = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_end.sizePolicy().hasHeightForWidth())
        self.line_end.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.line_end.setFont(font)
        self.line_end.setObjectName("line_end")
        self.gridLayout.addWidget(self.line_end, 2, 5, 1, 1)
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.gridLayout.addWidget(self.label_7, 2, 6, 1, 1)
        self.button_unlink = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_unlink.setFont(font)
        self.button_unlink.setObjectName("button_unlink")
        self.gridLayout.addWidget(self.button_unlink, 3, 11, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_8.setFont(font)
        self.label_8.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_8.setObjectName("label_8")
        self.gridLayout.addWidget(self.label_8, 2, 8, 1, 1)
        self.button_output = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_output.setFont(font)
        self.button_output.setObjectName("button_output")
        self.gridLayout.addWidget(self.button_output, 4, 6, 1, 1)
        self.button_rlsol = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_rlsol.setFont(font)
        self.button_rlsol.setObjectName("button_rlsol")
        self.gridLayout.addWidget(self.button_rlsol, 2, 11, 1, 1)
        self.line_lambda = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_lambda.sizePolicy().hasHeightForWidth())
        self.line_lambda.setSizePolicy(sizePolicy)
        self.line_lambda.setObjectName("line_lambda")
        self.gridLayout.addWidget(self.line_lambda, 1, 10, 1, 1)
        self.drop_method = QtWidgets.QComboBox(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.drop_method.sizePolicy().hasHeightForWidth())
        self.drop_method.setSizePolicy(sizePolicy)
        self.drop_method.setObjectName("drop_method")
        self.gridLayout.addWidget(self.drop_method, 1, 1, 1, 1)
        self.label_15 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_15.setFont(font)
        self.label_15.setFrameShape(QtWidgets.QFrame.Box)
        self.label_15.setFrameShadow(QtWidgets.QFrame.Raised)
        self.label_15.setLineWidth(3)
        self.label_15.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.label_15.setObjectName("label_15")
        self.gridLayout.addWidget(self.label_15, 0, 0, 1, 12)
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_5.sizePolicy().hasHeightForWidth())
        self.label_5.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_5.setFont(font)
        self.label_5.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 2, 0, 1, 2)
        self.label_9 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_9.sizePolicy().hasHeightForWidth())
        self.label_9.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.gridLayout.addWidget(self.label_9, 3, 0, 1, 2)
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_10.sizePolicy().hasHeightForWidth())
        self.label_10.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.gridLayout.addWidget(self.label_10, 4, 0, 1, 2)
        self.wid_date = QtWidgets.QDateEdit(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.wid_date.setDate(QtCore.QDate.currentDate())
        self.wid_date.setDisplayFormat("dd-MMM-yyyy")
        self.wid_date.setFont(font)
        self.wid_date.setObjectName("wid_date")
        self.gridLayout.addWidget(self.wid_date, 2, 2, 1, 1)
        self.line_dat = QtWidgets.QLineEdit(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.line_dat.setFont(font)
        self.line_dat.setObjectName("line_dat")
        self.gridLayout.addWidget(self.line_dat, 3, 2, 1, 4)
        self.line_output = QtWidgets.QLineEdit(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.line_output.setFont(font)
        self.line_output.setObjectName("line_output")
        self.gridLayout.addWidget(self.line_output, 4, 2, 1, 4)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setVerticalSpacing(15)
        self.gridLayout_2.setObjectName("gridLayout_2")
        spacerItem = QtWidgets.QSpacerItem(10, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem, 0, 5, 1, 1)
        self.label_11 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_11.sizePolicy().hasHeightForWidth())
        self.label_11.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_11.setFont(font)
        self.label_11.setAlignment(QtCore.Qt.AlignBottom|QtCore.Qt.AlignHCenter)
        self.label_11.setObjectName("label_11")
        self.gridLayout_2.addWidget(self.label_11, 0, 0, 1, 2)
        self.label_13 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_13.sizePolicy().hasHeightForWidth())
        self.label_13.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_13.setFont(font)
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.gridLayout_2.addWidget(self.label_13, 0, 6, 1, 1)
        self.button_del = QtWidgets.QPushButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.button_del.sizePolicy().hasHeightForWidth())
        self.button_del.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_del.setFont(font)
        self.button_del.setObjectName("button_del")
        self.gridLayout_2.addWidget(self.button_del, 4, 0, 1, 1)
        self.button_log = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_log.setFont(font)
        self.button_log.setObjectName("button_log")
        self.gridLayout_2.addWidget(self.button_log, 4, 6, 1, 1)
        self.label_14 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_14.setFont(font)
        self.label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.label_14.setObjectName("label_14")
        self.gridLayout_2.addWidget(self.label_14, 2, 6, 1, 1)
        self.label_12 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_12.sizePolicy().hasHeightForWidth())
        self.label_12.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.label_12.setFont(font)
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.gridLayout_2.addWidget(self.label_12, 0, 3, 1, 2)
        self.list_signals = QtWidgets.QListWidget(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.list_signals.sizePolicy().hasHeightForWidth())
        self.list_signals.setSizePolicy(sizePolicy)
        self.list_signals.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.list_signals.setFont(font)
        self.list_signals.setObjectName("list_signals")
        self.gridLayout_2.addWidget(self.list_signals, 1, 3, 3, 2)
        self.button_del_all = QtWidgets.QPushButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.button_del_all.sizePolicy().hasHeightForWidth())
        self.button_del_all.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.button_del_all.setFont(font)
        self.button_del_all.setObjectName("button_del_all")
        self.gridLayout_2.addWidget(self.button_del_all, 4, 1, 1, 1)
        self.list_dat = QtWidgets.QListWidget(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.list_dat.sizePolicy().hasHeightForWidth())
        self.list_dat.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.list_dat.setFont(font)
        self.list_dat.setObjectName("list_dat")
        self.gridLayout_2.addWidget(self.list_dat, 1, 0, 3, 2)
        self.text_log = QtWidgets.QTextEdit(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.text_log.setFont(font)
        self.text_log.setObjectName("text_log")
        self.gridLayout_2.addWidget(self.text_log, 3, 6, 1, 1)
        self.wid_osc = PlotWidget(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.wid_osc.sizePolicy().hasHeightForWidth())
        self.wid_osc.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Bosch Office Sans")
        font.setPointSize(10)
        self.wid_osc.setFont(font)
        self.wid_osc.setObjectName("wid_osc")
        self.gridLayout_2.addWidget(self.wid_osc, 1, 6, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(10, 20, QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem1, 0, 2, 1, 1)
        self.verticalLayout_2.addLayout(self.gridLayout_2)
        self.gridLayout.addLayout(self.verticalLayout_2, 6, 0, 1, 12)
        self.line_up = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_up.sizePolicy().hasHeightForWidth())
        self.line_up.setSizePolicy(sizePolicy)
        self.line_up.setObjectName("line_up")
        self.gridLayout.addWidget(self.line_up, 1, 7, 1, 1)
        self.line_down = QtWidgets.QLineEdit(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.line_down.sizePolicy().hasHeightForWidth())
        self.line_down.setSizePolicy(sizePolicy)
        self.line_down.setObjectName("line_down")
        self.gridLayout.addWidget(self.line_down, 2, 7, 1, 1)
        self.verticalLayout.addLayout(self.gridLayout)
        Transient_Compilation.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(Transient_Compilation)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1101, 21))
        self.menubar.setObjectName("menubar")
        self.menuOpen = QtWidgets.QMenu(self.menubar)
        self.menuOpen.setObjectName("menuOpen")
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        Transient_Compilation.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(Transient_Compilation)
        self.statusbar.setObjectName("statusbar")
        Transient_Compilation.setStatusBar(self.statusbar)
        self.actionOpen_Measurements = QtWidgets.QAction(Transient_Compilation)
        self.actionOpen_Measurements.setObjectName("actionOpen_Measurements")
        self.actionExit_Application = QtWidgets.QAction(Transient_Compilation)
        self.actionExit_Application.setObjectName("actionExit_Application")
        self.menuOpen.addAction(self.actionOpen_Measurements)
        self.menuOpen.addSeparator()
        self.menuOpen.addAction(self.actionExit_Application)
        self.menubar.addAction(self.menuOpen.menuAction())
        self.menubar.addAction(self.menuAbout.menuAction())

        self.retranslateUi(Transient_Compilation)
        QtCore.QMetaObject.connectSlotsByName(Transient_Compilation)

        self.lambda_drop()
        self.rlsol_drop()
        self.caln()
        self.method()
    def retranslateUi(self, Transient_Compilation):
        _translate = QtCore.QCoreApplication.translate
        Transient_Compilation.setWindowTitle(_translate("Transient_Compilation", "Transient Compilation"))
        self.label_16.setText(_translate("Transient_Compilation", "Method"))
        self.button_link.setText(_translate("Transient_Compilation", "Link"))
        self.rb_move.setText(_translate("Transient_Compilation", "Move Files "))
        self.rb_copy.setText(_translate("Transient_Compilation", "Copy Files"))
        self.button_output_folder.setText(_translate("Transient_Compilation", "Output Folder"))
        self.label.setText(_translate("Transient_Compilation", "Project Name"))
        self.label_2.setText(_translate("Transient_Compilation", "Start Time (in Sec) :"))
        self.label_3.setText(_translate("Transient_Compilation", "Lean Dev (in %) (Acc,Dec)"))
        self.button_lambda.setText(_translate("Transient_Compilation", "Add"))
        self.label_6.setText(_translate("Transient_Compilation", "End Time (in Sec):"))
        self.label_4.setText(_translate("Transient_Compilation", "Lambda"))
        self.button_reset.setText(_translate("Transient_Compilation", "Reset UI"))
        self.button_compile.setText(_translate("Transient_Compilation", "Compile"))
        self.button_dat.setText(_translate("Transient_Compilation", "Browse"))
        self.label_7.setText(_translate("Transient_Compilation", "Rich Dev (in %) (Acc,Dec)"))
        self.button_unlink.setText(_translate("Transient_Compilation", "UnLink"))
        self.label_8.setText(_translate("Transient_Compilation", "rlsol"))
        self.button_output.setText(_translate("Transient_Compilation", "Browse"))
        self.button_rlsol.setText(_translate("Transient_Compilation", "Add"))
        self.label_15.setText(_translate("Transient_Compilation", "TRANSIENT COMPILATION TOOL"))
        self.label_5.setText(_translate("Transient_Compilation", "Date"))
        self.label_9.setText(_translate("Transient_Compilation", "Measurement Files"))
        self.label_10.setText(_translate("Transient_Compilation", "Output Folder"))
        self.label_11.setText(_translate("Transient_Compilation", "List of Loaded Files"))
        self.label_13.setText(_translate("Transient_Compilation", "Signals Selected"))
        self.button_del.setText(_translate("Transient_Compilation", "Delete"))
        self.button_log.setText(_translate("Transient_Compilation", "Clear Log"))
        self.label_14.setText(_translate("Transient_Compilation", "Activity Log"))
        self.label_12.setText(_translate("Transient_Compilation", "List of signlas"))
        self.button_del_all.setText(_translate("Transient_Compilation", "Delete All"))
        self.menuOpen.setTitle(_translate("Transient_Compilation", "Open"))
        self.menuAbout.setTitle(_translate("Transient_Compilation", "About"))
        self.actionOpen_Measurements.setText(_translate("Transient_Compilation", "Open Measurements"))
        self.actionExit_Application.setText(_translate("Transient_Compilation", "Exit Application"))
        #ShortCuts
        self.actionOpen_Measurements.setShortcut("Ctrl+O")
        self.actionExit_Application.setShortcut("Alt+F4")
        self.line_output.setText(self.output_folder[self.user])
        self.line_output.setReadOnly(True)
        self.line_dat.setReadOnly(True)
        #clicked
        self.actionOpen_Measurements.triggered.connect(self.dat_open)
        self.button_dat.pressed.connect(self.dat_open)
        # self.drop_method.addItems(self.method)
        self.drop_method.setCurrentIndex(1)
        self.button_output.pressed.connect(self.out_folder)
        self.button_lambda.pressed.connect(self.lambda_add)
        self.button_rlsol.pressed.connect(self.rlsol_add)
        self.rb_copy.setChecked(True)
        self.list_dat.itemClicked.connect(self.signal_list)
        self.button_output_folder.pressed.connect(self.out_folder_pop)
        self.button_link.pressed.connect(self.link)
        self.button_compile.pressed.connect(self.compile)
        self.button_c = QtWidgets.QPushButton(self.wid_osc)
        self.button_c.setText('C')
        self.button_c.setFixedSize(20,20)
        self.wid_osc.setBackground(self.color_white)
        self.wid_osc.showGrid(x=True, y=True)
        self.wid_osc.setLabel('bottom', 'Time (in sec)', color='Blue', size=20)
        self.wid_osc.setLabel('left', 'Label Value', color='Blue', size=20)
        self.button_c.pressed.connect(self.clear_osc)
        self.list_signals.itemDoubleClicked.connect(self.graph_plot)
        self.button_unlink.pressed.connect(self.unlink)
        self.button_del.pressed.connect(self.Del_Mod)
        self.button_del_all.pressed.connect(self.Del_all)
        self.actionExit_Application.triggered.connect(self.close_application)
        self.text_log.setReadOnly(True)
        self.text_log.append('Welcome to Transient Compilation Tool')
        self.text_log.append('Please add measurement files to start.')
        self.button_log.pressed.connect(self.clr_log)
        self.button_reset.pressed.connect(self.reset_ui)
    if True:
        user = 'User'
        default_dir = {
        'Ritesh' : 'C:/Users/IIH3KOR/Desktop/YP8_CNG_BS6_TRANSIENTS/Task 03012020/YP8_CNG_Transient/CNG_Transient',
        'User'   : 'C:/Users'
        }
        template_file = {
        'Ritesh' : 'C:/Users/IIH3KOR/Documents/Python Automation/Kundan requirement/YBA load_change/template_file.xlsx',
        'User'   : 'C:/BOSCH Motronic Tools/Transient_Compilation/Template/template_file.xlsx'
        }
        output_folder = {
        'Ritesh' : 'C:/Users/IIH3KOR/Documents/Python Automation/Kundan requirement/YBA load_change/Output',
        'User'   : 'C:/BOSCH Motronic Tools/Transient_Compilation/Output'
        }
        alias_names = {
        'Ritesh' : 'C:/Users/IIH3KOR/Documents/Python Automation/Kundan requirement/YBA load_change/Alias_Names.xlsx',
        'User'   : 'C:/BOSCH Motronic Tools/Transient_Compilation/Alias Names/Alias_Names.xlsx'
        }
        color_dat = QtGui.QColor("#E5E4E2")
        color_white = QtGui.QColor("#FFFFFF")
        color_signals = QtGui.QColor("#ADD8E6")
        color_missing = QtGui.QColor("#ff4d4d")
        color_linked = QtGui.QColor("#ff8533")
        color_compiled = QtGui.QColor("#33cc00")
        dat_files = []
        lambda_list=[]
        rlsol_list=[]
        method_list = []
        keys = ['file','up','down','start_time','end_time','project','method']
        data = {}
        for key in keys:
            data[key] = []
    def dat_open(self) :
        if self.line_dat.text() is '' :
            open = self.default_dir[self.user]
        else :
            open = str(self.line_dat.text())
        name = QtWidgets.QFileDialog.getOpenFileNames(None, 'Measurement Files', open ,"Measurement Files (*.dat *.mf4)")
        items = [(self.list_dat.item(index).text()) for index in range(self.list_dat.count()) ]
        duplicate = 0
        added = 0
        for item in name[0] :
            if (os.path.basename(item) in items) :
                duplicate += 1
                continue
            added += 1
            QtWidgets.QListWidgetItem(os.path.basename(item),self.list_dat).setBackground(self.color_dat)    #After Selecting BG color changes to Red by default
            self.dat_files.append(item)
            self.line_dat.setText(os.path.dirname(item))
        if added :
            self.text_log.append(str(added) + ' Measurements added.')
        if duplicate >= 1 :
            QtWidgets.QMessageBox.question(None, 'Alert',str(duplicate) + " File already Present", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Ok))
    def caln(self) :
        self.wid_date.setCalendarPopup(True)
    def out_folder(self) :
        if self.line_output.text() is not self.output_folder[self.user]:
            open = str(self.line_output.text())
        else:
            open = self.output_folder[self.user]
        name = QtWidgets.QFileDialog.getExistingDirectory(None, 'Select Folder',open)
        if len(name) is 0 :
            self.line_output.setText(self.output_folder[self.user])
        else:
            self.line_output.setText(name)
    def lambda_add(self) :
        alias_excel = openpyxl.load_workbook(self.alias_names[self.user])
        alias_sheet = alias_excel["Alias_Names"]
        len_a_lambda = [i for i in range(1,100) if alias_sheet['D'+str(i)].value is not None][-1] + 1
        check_list = list(alias_sheet['D'])
        check = True
        for i in range(1,len_a_lambda):
             if self.line_lambda.text() == check_list[i].value or self.line_lambda.text() is '':
                 check = False
                 break
        if check :
            alias_sheet.cell(row = len_a_lambda, column = 4).value = self.line_lambda.text()
            self.drop_lambda.addItem(self.line_lambda.text())
        alias_excel.save(self.alias_names[self.user])
    def close_application(self):
        choice = QtWidgets.QMessageBox.question(None, 'Quit!',"Are you sure you want to quit the Application", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Yes|QtWidgets.QMessageBox.No))
        if choice == QtWidgets.QMessageBox.Yes:
            sys.exit()
        else:
            pass
    def rlsol_add(self) :
        alias_excel = openpyxl.load_workbook(self.alias_names[self.user])
        alias_sheet = alias_excel["Alias_Names"]
        len_a_lambda = [i for i in range(1,100) if alias_sheet['C'+str(i)].value is not None][-1] + 1
        check_list = list(alias_sheet['C'])
        check = True
        for i in range(1,len_a_lambda):
             if self.line_rlsol.text() == check_list[i].value or self.line_rlsol.text() is '':
                 check = False
                 break
        if check :
            alias_sheet.cell(row = len_a_lambda, column = 3).value = self.line_rlsol.text()
            self.drop_rlsol.addItem(self.line_rlsol.text())
        alias_excel.save(self.alias_names[self.user])
    def lambda_drop(self) :
        alias_excel = openpyxl.load_workbook(self.alias_names[self.user])
        alias_st = alias_excel["Alias_Names"]
        len_lm = [i for i in range(1,100) if alias_st['D'+str(i)].value is not None][-1]
        check_list = list(alias_st['D'])
        for i in range(1, len_lm):
            self.lambda_list.append(check_list[i].value)
        self.drop_lambda.addItems(self.lambda_list)
    def rlsol_drop(self) :
        alias_excel = openpyxl.load_workbook(self.alias_names[self.user])
        alias_st = alias_excel["Alias_Names"]
        len_lm = [i for i in range(1,100) if alias_st['C'+str(i)].value is not None][-1]
        check_list = list(alias_st['C'])
        for i in range(1, len_lm):
            self.rlsol_list.append(check_list[i].value)
        self.drop_rlsol.addItems(self.rlsol_list)
    def method(self) :
        alias_excel = openpyxl.load_workbook(self.alias_names[self.user])
        alias_st = alias_excel["Method_List"]
        len_lm = [i for i in range(1,100) if alias_st['B'+str(i)].value is not None][-1]
        check_list = list(alias_st['B'])
        for i in range(1, len_lm):
            self.method_list.append(check_list[i].value)
        self.drop_method.addItems(self.method_list)
    def signal_list(self) :
        selected_dat_file = self.dat_files[self.list_dat.currentRow()]
        self.line_dat.setText(os.path.dirname(selected_dat_file))
        if self.already_linked(selected_dat_file) :
            self.set_dropdowns(selected_dat_file)
        self.list_signals.clear()
        self.wid_osc.clear()
        self.list_signals.setSortingEnabled(True)
        signals = mdfreader.MdfInfo().list_channels(self.dat_files[self.list_dat.currentRow()])
        for signal in signals :
            if signal.split('$')[0] and signal.split('_')[0] is not '' :
                items = [(self.list_signals.item(index).text()) for index in range(self.list_signals.count()) ]
                if signal not in items and signal != 'time' :
                    QtWidgets.QListWidgetItem(signal,self.list_signals).setBackground(self.color_white)
    def out_folder_pop(self) :
        os.startfile(self.line_output.text())
    def link(self) :
        if self.list_dat.currentRow() is not -1:
            def run_this() :
                properties = [self.line_pn,self.line_start,self.line_end]
                count = 0
                for i in range(len(properties)):
                    if properties[i].text() == '' :
                        properties[i].setStyleSheet('QLineEdit { background-color: '+'#ff4d4d'+';}')
                        count = count + 1
                    else :
                        properties[i].setStyleSheet('QLineEdit { background-color: '+'#FFFFFF'+';}')
                        continue
                if count :
                    self.text_log.append('Please fill the missing details to link')
                if count == 0 :
                    for i in range(len(properties)):
                        properties[i].setStyleSheet('QLineEdit { background-color: '+'#FFFFFF'+';}')
                    self.append_data_to_compile()
                    self.list_dat.currentItem().setBackground(self.color_linked)
                    self.text_log.append('File Linked')
            if self.already_linked(self.dat_files[self.list_dat.currentRow()]) :
                choice = QtWidgets.QMessageBox.question(None, 'Alert' , " It is already link. Do you want to overwrite?", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Yes|QtWidgets.QMessageBox.No))
                if choice == QtWidgets.QMessageBox.Yes:
                    self.unlink_run_this(self.dat_files[self.list_dat.currentRow()])
                    run_this()
                else :
                    self.set_dropdowns(self.dat_files[self.list_dat.currentRow()])
            else :
                    run_this()
    def already_linked(self, file):
        if (file in self.data['file'] ):
            return True
        return False
    def all_files_linked(self):
        if self.list_dat.count() == len(self.data['file']):
            return True
        else :
            QtWidgets.QMessageBox.question(None, 'Warning!',"Please link all the files before compiling", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Ok))
        return False
    def compile(self) :
        if self.all_files_linked():
            try:
                if self.rb_copy.isChecked() :
                    folder_action = 'Copy'
                else :
                    folder_action = 'Move'
                Transientv203.compile(self.data,self.line_output.text(),folder_action,self.template_file[self.user],self.alias_names[self.user])
                if self.rb_copy.isChecked() :
                    self.text_log.append('Copied Files')
                else :
                    self.text_log.append('Moved Files')
                for i in range(self.list_dat.count()):
                    self.list_dat.item(i).setBackground(self.color_compiled)
                self.text_log.append('Compiled')
                QtWidgets.QMessageBox.information(self.centralwidget, 'Done!', 'Compilation Successfull!', QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Ok))
                self.text_log.append('Please Reset UI to use it again.')
            except Exception as e:
                self.text_log.append('Please contact tool developer Ritesh because its a '+str(e)+' error.')
    def append_data_to_compile(self):
        self.data['file'].append(self.dat_files[self.list_dat.currentRow()])
        if self.line_up.text() != '' :
            self.data['up'].append(self.line_up.text())
        else :
            self.data['up'].append('100')
        if self.line_down.text() != '' :
            self.data['down'].append(self.line_down.text())
        else :
            self.data['down'].append('100')
        self.data['start_time'].append(self.line_start.text())
        self.data['end_time'].append(self.line_end.text())
        self.data['project'].append(self.line_pn.text())
        self.data['method'].append(self.drop_method.currentText())
    def graph_plot(self) :
        try :
            if self.list_dat.currentRow() is not -1:
                mdfData=mdfreader.Mdf(self.dat_files[self.list_dat.currentRow()])
                mdfData.resample(1)
                selected_signal=mdfData.get_channel_data(self.list_signals.item(self.list_signals.currentRow()).text())
                self.draw(selected_signal, str(self.list_signals.item(self.list_signals.currentRow()).text()))
        except Exception as e:
            self.text_log.append(str(e))
    def clear_osc(self) :
        try:
            for index in range(self.list_signals.count()):
                self.list_signals.item(index).setBackground(self.color_white)
            self.wid_osc.clear()
        except Exception as e:
            self.text_log.append('What else do you want to clear!!')
    def draw(self, y, plotname) :
        color = ["#"+''.join([choice('0123456789ABCDEF') for j in range(6)]) for i in range(len(y))]
        selected_color = choice(color)
        self.list_signals.currentItem().setBackground(QtGui.QColor(selected_color))
        self.wid_osc.plot(y, name=plotname, pen=mkPen(color=str(selected_color)), symbol='t', symbolSize=5, symbolBrush=str(selected_color))
    def set_dropdowns(self, file):
        index = self.data['file'].index(file)
        self.drop_method.setCurrentText(self.data['method'][index])
        self.line_pn.setText(self.data['project'][index])
        self.line_start.setText(self.data['start_time'][index])
        self.line_end.setText(self.data['end_time'][index])
        self.line_up.setText(self.data['up'][index])
        self.line_down.setText(self.data['down'][index])
    def unlink(self) :
        if self.list_dat.currentRow() is not -1 :
            selected_dat_file = self.dat_files[self.list_dat.currentRow()]
            if self.already_linked(selected_dat_file) :
                choice = QtWidgets.QMessageBox.question(None, 'Unlink',"Are you sure you want to unlink", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Yes|QtWidgets.QMessageBox.No))
                if choice == QtWidgets.QMessageBox.Yes :
                    self.unlink_run_this(selected_dat_file)
                    self.wid_osc.clear()
                    self.list_signals.clear()
                    return(True)
                else :
                    return(False)
            else :
                self.wid_osc.clear()
                self.list_signals.clear()
                return(True)
    def unlink_run_this(self,file):
        index = self.data['file'].index(file)
        self.list_dat.currentItem().setBackground(self.color_dat)
        for key in self.keys:
            self.data[key].pop(index)
    def Del_Mod(self) :
        self.unlink()
        if self.unlink() :
            self.dat_files.pop(self.list_dat.currentRow())
            self.list_dat.takeItem(self.list_dat.currentRow())
            self.text_log.append('Measurement File Removed')
    def Del_all(self) :
        self.dat_files.clear()
        self.line_dat.setText('')
        self.list_dat.clear()
        self.wid_osc.clear()
        self.list_signals.clear()
        for key in self.keys:
            self.data[key] = []
    def clr_log(self) :
        self.text_log.clear()
        self.text_log.append('Welcome to Transient Compilation Tool')
        self.text_log.append('Please add measurement files to start.')
        QtWidgets.QMessageBox.information(None, 'Alert!',"Log Cleared", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Ok))
    def reset_ui(self) :
         self.reset_details()
         self.Del_all()
         self.text_log.clear()
         for key in self.keys:
             self.data[key] = []
         self.text_log.append('Welcome Back')
         QtWidgets.QMessageBox.question(None, 'Alert!',"Reset Successfull", QtWidgets.QMessageBox.StandardButtons(QtWidgets.QMessageBox.Ok))
    def reset_details(self) :
        self.drop_method.setCurrentIndex(1)
        self.wid_date.setDate(QtCore.QDate.currentDate())
        self.line_pn.setText('')
        self.line_start.setText('')
        self.line_end.setText('')
        self.line_up.setText('')
        self.line_down.setText('')
        self.line_rlsol.setText('')
        self.line_lambda.setText('')
        self.line_dat.setText('')
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Transient_Compilation = QtWidgets.QMainWindow()
    ui = Ui_Transient_Compilation()
    ui.setupUi(Transient_Compilation)
    Transient_Compilation.show()
    sys.exit(app.exec_())
